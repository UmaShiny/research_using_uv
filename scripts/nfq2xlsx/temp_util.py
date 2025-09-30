import openpyxl as pl
from pathlib import Path
import openpyxl.worksheet.worksheet as worksheet


def set_sheet_title():
    comp_path = Path(".\\data\\temp_xlsx\\temp_xlsx_complete")
    comp_files = [f for f in comp_path.glob("*")]

    for file in comp_files:
        print(f"processing: {file.name}")
        wb = pl.load_workbook(filename=file)
        ws = wb.active
        new_title = f"temp_char {file.name.split('.xlsx')[0]}"
        if not isinstance(ws, worksheet.Worksheet):
            raise TypeError("Failed to load worksheet.")
        ws.title = new_title
        wb.save(file)
        print(f"change title: {file.name} to {new_title}")


def modify_column():
    comp_path = Path(".\\data\\temp_xlsx\\temp_xlsx_complete")
    comp_files = [f for f in comp_path.glob("*")]

    for file in comp_files:
        print(f"processing: {file.name}")
        wb = pl.load_workbook(filename=file)
        ws = wb.active
        if not isinstance(ws, worksheet.Worksheet):
            raise TypeError("Failed to load worksheet.")
        ws.cell(row=1, column=1, value="決算発表月")
        ws.cell(row=2, column=1, value="決算発表日")
        wb.save(file)
        print(f"modify column: {file.name} A1 to 決算発表月, A2 to 決算発表日")


modify_column()
