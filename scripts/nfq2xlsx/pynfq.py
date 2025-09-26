import os as _os
from pathlib import Path as _Path
import win32com.client as _com
import re as _re
import pandas as _pd
from typing import Optional, Union, overload, Literal
import openpyxl as _xl
from openpyxl.cell.cell import MergedCell, Cell
from openpyxl.worksheet.worksheet import Worksheet as _ws
import sys
from datetime import datetime as dt
import numpy as np

# Use for important matters when displaying in the prompt.
RED = "\033[31m"
GRN = "\033[32m"
END = "\033[0m"


# define helper function
def find_all_file_from_dir(src_dirpath: str, mode: str) -> list[str]:
    folder = _Path(src_dirpath)
    all_file: list[str] = []
    if mode == "final":
        all_file = [f.name for f in folder.rglob("*") if f.is_file()]
    if mode == "absol":
        all_file = [str(f) for f in folder.rglob("*") if f.is_file()]
    return all_file


def extract_dst_file_name(sheetName: str, suffix: str) -> str:
    searched_Ticker = _re.search(pattern=r"T\d{3}(\d|A)", string=sheetName)
    if searched_Ticker is None:
        raise TypeError(
            f"Unexpected Excel worksheet name. What was requested was *[searched_Ticker].xlsx : sheetName = {sheetName}"
        )
    dst_file_name = f"{searched_Ticker.group()}{suffix}"
    return dst_file_name


def validate_arg_type(
    arg: Union[object, list[object], None],
    required_type: Union[list[type], list[Optional[type]], type, None],
    arg_name: Union[str, list[str]],
) -> None:
    # Case.1 : arg or arg_name is list
    if isinstance(arg, list) or isinstance(arg_name, list):
        if (isinstance(arg, list) and isinstance(arg_name, str)) or (
            not isinstance(arg, list) and isinstance(arg_name, list)
        ):
            raise ValueError(
                f"Type mismatch: {type(arg).__name__} != {type(arg_name).__name__}"
            )
        elif isinstance(arg, list) and isinstance(arg_name, list):
            if len(arg) == len(arg_name):
                for arg_ele, arg_name_ele in list(zip(arg, arg_name)):
                    validate_arg_type(
                        arg=arg_ele, required_type=required_type, arg_name=arg_name_ele
                    )
            else:
                raise ValueError(
                    f"Length mismatch: len(arg)={len(arg)}, len(arg_name)={len(arg_name)}"
                )

    # Case.2 : required_type is None
    elif required_type is None:
        if arg is None:
            return
        else:
            raise ValueError(
                f"Type mismatch: {type(arg).__name__} != {type(arg_name).__name__}"
            )

    # Case.3 : required_type is list
    else:
        if isinstance(required_type, list):
            for required_type_ele in required_type:
                if required_type_ele is None:
                    required_type_ele = type(required_type_ele)
                if isinstance(arg, required_type_ele):
                    return
            raise ValueError(
                f"{arg_name} is not {" ,".join([type(None).__name__ if required_type_ele is None else required_type_ele.__name__ for required_type_ele in required_type])}. {arg_name}={type(arg).__name__}"
            )

        # Case.4 : required_type is one type
        elif isinstance(required_type, type) and isinstance(arg, required_type):
            return

        raise ValueError(
            f"{arg_name} is not {required_type.__name__}. {arg_name}={type(arg).__name__}"
        )


def validate_path(path: str, var_name: str) -> None:
    if _os.path.exists(path):
        raise ValueError(f"This path is already exists\n{var_name} = {path}")


def validate_isdir(
    dir_path: Union[str, list[str]], var_name: Union[str, list[str]]
) -> None:
    if isinstance(dir_path, str) and isinstance(var_name, str):
        if not _os.path.isdir(dir_path):
            raise ValueError(f"src_dirpath is not a folder : {var_name} = {dir_path}")
    elif isinstance(dir_path, list) and isinstance(var_name, list):
        if len(dir_path) == len(var_name):
            for dir_path_ele, var_name_ele in list(zip(dir_path, var_name)):
                validate_isdir(dir_path=dir_path_ele, var_name=var_name_ele)
        else:
            raise ValueError(
                f"Length mismatch: len(dir_path)={len(dir_path)}, len(var_name)={len(var_name)}"
            )


def print_progress(current: int, total: int, bar_length: int = 30) -> None:

    def _helper_display(display: str) -> None:
        sys.stdout.write(display)
        sys.stdout.flush()

    fraction = current / total
    filled_length = int(bar_length * fraction)
    bar = "█" * filled_length + "-" * (bar_length - filled_length)
    percent = fraction * 100
    display = f"[{bar}] {current}/{total} ({percent:.1f}%)"

    if current < total:
        _helper_display(f"\r{display}")
    elif current == total:
        _helper_display(f"\r{GRN}{display} complete{END}\n")


# define class about handle *.xlsx from nfq and converse to *.csv through _pd.DataFrame format
class pyNFQ:

    # When using the financial results announcement date as an attribute, define the format in which the value is described.
    # note: FRAD = "Financial Results Announcement Date"
    FRAD_FORMAT = "%Y/%m/%d"

    # Absolute path of the currently running Python file
    script_file_dir: str | None = None

    # If there is a CSV file containing the target tickers, store it as a DataFrame.
    Tickers: _pd.DataFrame | None = None

    # Maintains the instance variable Tickers in list format
    Tickers_list: list | None = None

    # Holds the number of tickers contained in the instance variable Tickers.
    Tickers_length: int | None = None

    # A variable used to fill in missing values on the financial results announcement date.
    # By using it as the difference from the first day of the fiscal month, it supplements the missing values.
    FR_FRA_tmstmp_diff: float | None = None

    # During initialization, register all tickers to be handled by this object.
    def __init__(
        self,
        path_to_ticker_symbol_csv: str | None,
        FR_FRA_tmstmp_diff: float | None = None,
    ) -> None:

        # Validation Zone
        validate_arg_type(
            arg=path_to_ticker_symbol_csv,
            required_type=[str, None],
            arg_name="path_to_ticker_symbol_csv",
        )

        # Initialize Instance Variance
        self.script_file_dir = _os.path.dirname(_os.path.abspath(__file__))
        self.Tickers = (
            None
            if path_to_ticker_symbol_csv is None
            else _pd.read_csv(path_to_ticker_symbol_csv)
        )
        self.Tickers_list = (
            None if self.Tickers is None else self.Tickers["Ticker"].tolist()
        )
        self.Tickers_length = (
            None if self.Tickers_list is None else len(self.Tickers_list)
        )
        self.FR_FRA_tmstmp_diff = FR_FRA_tmstmp_diff

    # Lightweight accessories
    def get_script_file_dir(self) -> str:
        if self.script_file_dir is None:
            raise ValueError(
                "The value for self.script_file_dir has not yet been assigned."
            )
        return self.script_file_dir

    def get_Tickers_list(self) -> Union[list, None]:
        return self.Tickers_list

    def get_Tickers_length(self) -> Union[int, None]:
        return self.Tickers_length

    def __helper_join_abspath(self, *args: str) -> str:
        dst = "\\".join(args)
        return _os.path.join(self.get_script_file_dir(), dst)

    # The following are method definitions.

    # Files downloaded from nfq cannot be used as xlsx files.
    # This function makes them usable as xlsx files.
    def repair_nfqxlsx(
        self,
        src_dirpath: str,
        dst_dirpath: str,
        folder_exist_check: bool = False,
        nfq_xlsx_overwrite: bool = False,
        validation: bool = False,
    ) -> None:
        """If you want to check the specified folder by setting folder_exist_check to True, you must set validation to True."""

        # This may have been left in an overwritable state due to user negligence, so it must be removed.
        if nfq_xlsx_overwrite:
            check = input(
                f"{RED}Overwriting is permitted, so the original file will be lost [y/n]>> {END}"
            )
            if check != "y":
                return print(
                    '"nfq_xlsx_overwrite" was set to "True", but execution was canceled following user confirmation.\nIf you do not permit overwriting, set it to False.'
                )

        # Validation Zone
        if validation:
            validate_arg_type(
                arg=[folder_exist_check, nfq_xlsx_overwrite],
                required_type=bool,
                arg_name=["folder_exist_check", "nfq_xlsx_overwrite"],
            )
            validate_arg_type(
                arg=[src_dirpath, dst_dirpath],
                required_type=str,
                arg_name=["src_dirpath", "dst_dirpath"],
            )
            validate_isdir(
                dir_path=[src_dirpath, dst_dirpath],
                var_name=["src_dirpath", "dst_dirpath"],
            )
            if folder_exist_check:
                validate_path(path=dst_dirpath, var_name="dst_dirpath")

        files = find_all_file_from_dir(src_dirpath, mode="absol")
        num_of_files = len(files)

        # Python use Excel.Application through COM
        for i, nfq_xlsx_file in enumerate(files):
            excel = _com.Dispatch("Excel.Application")
            print_progress(current=i + 1, total=num_of_files, bar_length=30)
            try:
                src_file_abspath = self.__helper_join_abspath(nfq_xlsx_file)
                wb = excel.Workbooks.Open(src_file_abspath)
                active_ws_title = wb.Sheets(2).Name
                excel.DisplayAlerts = False
                wb.Sheets(1).Delete()  # Since "検索条件シート" is no longer needed,
                # removing it here will reduce processing overhead going forward.
                excel.DisplayAlerts = True
                if nfq_xlsx_overwrite:
                    wb.Save()
                else:
                    dst_file_abspath = self.__helper_join_abspath(
                        dst_dirpath, extract_dst_file_name(active_ws_title, ".xlsx")
                    )
                    wb.SaveAs(dst_file_abspath)
                wb.Close()
            except Exception as e:
                print(f"Error Occured : {e} in {self.repair_nfqxlsx.__name__}")
            finally:
                excel.Quit()

    # Function for handling standard *.xlsx files with _pd.DataFrame.
    # Performs header/index removal and missing value imputation for earnings release dates.
    def xlsx2csv(
        self,
        src_dirpath: str,
        dst_dirpath: str,
        remove_column: list[int],
        remove_row: list[int],
        missing_replace: int | float = 0,
        folder_exist_check: bool = False,
        validation: bool = False,
    ) -> None:
        """If you want to check the specified folder by setting folder_exist_check to True, you must set validation to True."""

        if validation:
            validate_arg_type(
                arg=[src_dirpath, dst_dirpath],
                required_type=[str, None],
                arg_name=["src_dirpath", "dst_dirpath"],
            )
            validate_arg_type(
                arg=[remove_column, remove_row],
                required_type=[list],
                arg_name=["remove_column", "remove_row"],
            )
            validate_arg_type(
                arg=folder_exist_check,
                required_type=bool,
                arg_name="folder_exist_check",
            )
            validate_isdir(
                dir_path=[src_dirpath, dst_dirpath],
                var_name=["src_dirpath", "dst_dirpath"],
            )
            if folder_exist_check:
                validate_path(path=dst_dirpath, var_name="dst_dirpath")

        src_abspath = self.__helper_join_abspath(src_dirpath)
        dst_abspath = self.__helper_join_abspath(dst_dirpath)

        all_xl_file_lst: list[str] = find_all_file_from_dir(src_abspath, mode="final")
        missing_only_xl_file_lst: list[str] = []
        mean_diff_lst: list[float] = []

        def __inner_helper_xlsx_shaper(wb: _xl.Workbook):
            ws = wb.worksheets[0]
            ticker = extract_dst_file_name(
                ws.title, suffix=""
            )  # it can extract only ticker
            ws.freeze_panes = "A1"
            for delete_col in sorted(remove_column, reverse=True):
                ws.delete_cols(delete_col)
            for delete_row in sorted(remove_row, reverse=True):
                ws.delete_rows(delete_row)
            for cell in ws[1][1 : ws.max_column + 1]:
                if isinstance(cell, Cell):
                    cell.value = f"{cell.value}/01"
                else:
                    raise TypeError(
                        f"Type error: value must be {Cell.__name__} (type passed: {type(cell).__name__})"
                    )
            ticker_row: list[str] = [
                ticker if col > 1 else "Ticker" for col in range(1, ws.max_column + 1)
            ]
            ws.append(ticker_row)
            return ws

        def __inner_helper_calc_mean_diff(
            list_1: list[float], list_2: list[float]
        ) -> float:
            arr1 = np.array(list_1, dtype=float)
            arr2 = np.array(list_2, dtype=float)
            mask = ~np.isnan(arr1) & ~np.isnan(arr2)  # 欠損値を除外
            if np.sum(mask) == 0:
                return 0.0
            return float(np.mean(arr2[mask] - arr1[mask]))

        def __inner_helper_get_xlsxrow(
            ws: _ws, row: int
        ) -> tuple[Cell | MergedCell, ...]:
            return ws[row][1 : ws.max_column + 1]

        def __inner_helper_str2timestamp_nfqver(ws: _ws, mode: str) -> list[float]:
            if mode == "d":
                return [
                    (
                        dt.timestamp(dt.strptime(str(cell.value), self.FRAD_FORMAT))
                        if cell.value != "-"
                        else np.nan
                    )
                    for cell in __inner_helper_get_xlsxrow(ws, 2)
                ]
            elif mode == "m":
                return [
                    dt.timestamp(dt.strptime(str(cell.value), self.FRAD_FORMAT))
                    for cell in __inner_helper_get_xlsxrow(ws, 1)
                ]
            else:
                raise ValueError(f'type is required "m" or "d". given {mode}')

        def __inner_helper_missing_impute(
            ws: _ws,
            missing_only: bool,
        ) -> None:
            """if FR_mo1st : type is "m"\nif FRA_dt : type is "d" """

            # Calculate timestamps for all values in the fiscal_month and list them with their respective first day as the timestamp.
            (
                FR_mo1st_tmstmp_lst,
                FRA_dt_tmstmp_lst,
            ) = __inner_helper_str2timestamp_nfqver(
                ws, mode="m"
            ), __inner_helper_str2timestamp_nfqver(
                ws, mode="d"
            )  # check_nan -> Convert missing values to NaN
            # FR  = Financial Results
            # FRA = Financial Results Announcement

            mean_diff: float  # Variable to compensate for missing values on the financial results announcement date

            # Calculate the average difference in timestamps from the first day of the fiscal year-end month to the fiscal year-end announcement date.
            if missing_only:
                if self.FR_FRA_tmstmp_diff is None:
                    raise ValueError(
                        "The value for self.FR_FRA_tmstmp_diff has not yet been assigned."
                    )
                mean_diff = self.FR_FRA_tmstmp_diff
            else:
                mean_diff = __inner_helper_calc_mean_diff(
                    FR_mo1st_tmstmp_lst, FRA_dt_tmstmp_lst
                )
                mean_diff_lst.append(
                    mean_diff
                )  # For stocks with missing values throughout the entire period, use the mean value from the mean_diff_list.

            mod_FRA_date_tmstmp_lst = [
                (
                    FR_mo1st_tmstmp_lst[idx] + mean_diff
                    if np.isnan(FRA_date_timestamp)
                    else FRA_dt_tmstmp_lst[idx]
                )
                for idx, FRA_date_timestamp in enumerate(FRA_dt_tmstmp_lst)
            ]

            for idx, cell in enumerate(__inner_helper_get_xlsxrow(ws, 2)):
                if isinstance(cell, MergedCell):
                    raise TypeError(
                        f"Type error: value must be {Cell.__name__} (type passed: {type(cell).__name__})"
                    )
                cell.value = mod_FRA_date_tmstmp_lst[idx]

            # The FR column will not be used going forward, so delete it here.
            ws.delete_rows(1)

        def __inner_helper_xl2csv(ws: _ws, missing_only: bool):
            __inner_helper_missing_impute(ws, missing_only)
            # By transposing(np.array(list).T), you can convert it into a format that can be handled by a DataFrame.
            arr = np.array(list(ws.values)).T
            arr[1:, 0] = (
                _pd.to_datetime(arr[1:, 0].astype(float), unit="s", utc=True)
                .tz_convert("Asia/Tokyo")
                .strftime(self.FRAD_FORMAT)
            )
            # Here, replace all missing values with 0, np.nan and so on.
            arr = arr.astype(object)
            arr[arr == "-"] = missing_replace  # default "missing_replace" is 0

            dtf = _pd.DataFrame(data=arr[1:, :], columns=arr[0, :])
            dtf.set_index(["Ticker", "決算発表日"], inplace=True, drop=True)
            dtf.index.name = "Date"
            csv_file_name = extract_dst_file_name(
                ws.title, ".csv"
            )  # Ticker = _search_ticker_from_wstitle(ws)
            dtf.to_csv(_os.path.join(dst_abspath, csv_file_name))

        def __inner_helper_load_wb(dir, file):
            return _xl.load_workbook(self.__helper_join_abspath(dir, file))

        @overload
        def __inner_helper_processing_flow(
            xl_file_lst: list[str],
            missing_only: Literal[False],
            pre_end_step: Optional[int] = None,
            total: Optional[int] = None,
        ) -> tuple[int, int]: ...

        @overload
        def __inner_helper_processing_flow(
            xl_file_lst: list[str],
            missing_only: Literal[True],
            pre_end_step: Optional[int] = None,
            total: Optional[int] = None,
        ) -> None: ...

        def __inner_helper_processing_flow(
            xl_file_lst: list[str],
            missing_only: bool,
            pre_end_step: Optional[int] = None,
            total: Optional[int] = None,
        ) -> Optional[tuple[int, int]]:

            if len(xl_file_lst) == 0:
                return

            start = 1 if pre_end_step is None else pre_end_step
            total = len(xl_file_lst) if total is None else total

            for current_iter, xlsx_file_name in enumerate(xl_file_lst, start=start):
                wb = __inner_helper_load_wb(src_abspath, xlsx_file_name)
                if missing_only:
                    ws = wb.worksheets[0]
                else:  # ws = __inner_helper_xlsx_shaper(wb)
                    ws = __inner_helper_xlsx_shaper(wb)
                    if all(
                        FRA_date == "-" for FRA_date in ws[2][1 : ws.max_column + 1]
                    ):  # In this case, it will not be reflected in the progress bar.
                        missing_only_xl_file_lst.append(xlsx_file_name)
                        wb.save(
                            self.__helper_join_abspath(dst_abspath, xlsx_file_name)
                        )  # -> 2nd phase
                        print(f"missing only : {xlsx_file_name}")
                        continue

                __inner_helper_xl2csv(ws, missing_only=missing_only)
                print_progress(current_iter, total)

                # If you haven't finished processing all the files yet
                if current_iter < total:
                    continue

                # When all files have been processed
                if self.FR_FRA_tmstmp_diff is None:
                    self.FR_FRA_tmstmp_diff = float(np.average(mean_diff_lst))
                progress_current = current_iter - len(missing_only_xl_file_lst)
                progress_total = total

                return (progress_current, progress_total)

        end_step, total = __inner_helper_processing_flow(
            all_xl_file_lst, missing_only=False
        )
        __inner_helper_processing_flow(
            missing_only_xl_file_lst,
            missing_only=True,
            pre_end_step=end_step,
            total=total,
        )

    def comp_rslt_Tkcr(self, dir: str):
        if self.Tickers is None:
            return

        Tickers = sorted(self.get_Tickers_list())  # type: ignore
        results = sorted(find_all_file_from_dir(src_dirpath=dir, mode="final"))

        error = []
        idx = None
        comp_over = False
        for result in results:
            if comp_over:
                break
            idx = 0 if idx is None else idx + 1
            while result != f"{Tickers[idx]}.csv":
                idx += 1
                error.append(f"{Tickers[idx]}.csv")
                if idx == len(Tickers):
                    comp_over = True
                    break

        for err in error:
            print(f"err: {err}")
