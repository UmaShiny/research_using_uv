import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet._write_only import WriteOnlyWorksheet
from openpyxl.cell.cell import Cell
import numpy as np
import pandas as pd
from pathlib import Path


import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

    # devide the 1 excel file into 4 excel files acording to financial-documents-type and settlement-of-accounts-type

class compact_xlsx:
    class _util:

        @staticmethod
        def devide_xlsx(file_path: Path) -> list[pd.DataFrame]:
            # pd.set_option('future.no_silent_downcasting', True)
            if file_path.suffix != ".xlsx":
                raise TypeError(f"file_path must be .xlsx file, but got {file_path.suffix}")
            
            dfs = []
            
            each_doc_type_ranges = [
                tanshin_tandoku_range_row := {"start": 2, "end": 487},
                tanshin_renketsu_range_row := {"start": 488, "end": 973},
                yuukahoukoku_tandoku_range_row := {"start": 974, "end": 1459},
                yuukahoukoku_renketsu_range_row := {"start": 1460, "end": 1945},
            ]

            wb:Workbook = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active

            if not isinstance(ws, Worksheet):
                raise TypeError("Unexpected type for worksheet")
            
            date_range_row = {"start": 8, "end": ws.max_row}

            for doc_type in each_doc_type_ranges:
                tmp = []
                data_labels = ws["G"][doc_type["start"]:doc_type["end"] + 1]
                date_index = ws[2][date_range_row["start"]:date_range_row["end"] + 1]
                for row in ws.iter_rows(min_row=doc_type["start"] + 1, max_row=doc_type["end"] + 1, min_col=9, max_col=ws.max_column):
                    data_tmp = [cell.value for cell in row if isinstance(cell, Cell)]
                    tmp.append(data_tmp)
                data = dict(zip([cell.value for cell in data_labels if isinstance(cell, Cell)], tmp))
                tmp_df = pd.DataFrame(
                        data=data,
                        index=[cell.value for cell in date_index if isinstance(cell, Cell)],
                    )
                nan_mask = tmp_df.iloc[:, 1:] == '-'
                tmp_df.iloc[:, 1:] = tmp_df.iloc[:, 1:].mask(nan_mask, np.nan)
                # tmp_df.iloc[:, 1:] = tmp_df.iloc[:, 1:].replace('-', np.nan)
                dfs.append(tmp_df)

            return dfs


        @staticmethod
        def merge_xlsx(dfs: list[pd.DataFrame]) -> pd.DataFrame:
            pd.set_option('mode.chained_assignment', None)  # 一時的に警告抑制
            tanshin_tandoku_df, tanshin_renketsu_df, yuukahoukoku_tandoku_df, yuukahoukoku_renketsu_df = dfs

            tanshin_df = tanshin_renketsu_df.combine_first(tanshin_tandoku_df)
            yuukahoukoku_df = yuukahoukoku_renketsu_df.combine_first(yuukahoukoku_tandoku_df)

            mask_dup_id = tanshin_df['決算発表日'] == yuukahoukoku_df['決算発表日']
            merged_df = yuukahoukoku_df[mask_dup_id].combine_first(tanshin_df[mask_dup_id])

            mask_merge_none = merged_df['決算発表日'].isna()
            mask_tanshin_none = tanshin_df['決算発表日'].isna()
            mask_yuukahoukoku_none = yuukahoukoku_df['決算発表日'].isna()

            # tanshin_df = tanshin_df.dropna(axis=1, how='all')
            # yuukahoukoku_df = yuukahoukoku_df.dropna(axis=1, how='all')

            merged_df = merged_df[~mask_merge_none]
            tanshin_df = tanshin_df[~mask_dup_id & ~mask_tanshin_none]
            yuukahoukoku_df = yuukahoukoku_df[~mask_dup_id & ~mask_yuukahoukoku_none]

            merged_df = pd.concat([merged_df, tanshin_df, yuukahoukoku_df])
            merged_df.sort_index(inplace=True)
            return merged_df.T

    @staticmethod
    def devide_xlsx(file_path: Path) -> list[pd.DataFrame]:
        all_xlsx = compact_xlsx._util.devide_xlsx(file_path)
        return all_xlsx

    @staticmethod
    def merge_xlsx(dfs: list[pd.DataFrame]) -> pd.DataFrame:
        merged_xlsx = compact_xlsx._util.merge_xlsx(dfs)
        return merged_xlsx


for path in Path("data/nfq/fix_nfqxlsx").rglob("*.xlsx"):
    print(f"processing: {path.name}")
    xlsx_dfs = compact_xlsx.devide_xlsx(path)
    result = compact_xlsx.merge_xlsx(xlsx_dfs)
    result.to_excel(Path("data/nfq/concat_nfqxlsx") / path.name, sheet_name=path.stem)