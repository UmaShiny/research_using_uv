from pathlib import Path
import openpyxl as pl
import openpyxl.worksheet.worksheet as worksheet
import openpyxl.workbook.workbook as workbook
from Config import Path as cp
from Config import Constant as cc
import numpy as np
import pandas as pd
from datetime import datetime as dt
import time

# 1. get list of data\temp_xlsx\temp_xlsx_date & get list of data\temp_xlsx\temp_xlsx_main
# 2. compare the head of list of file-name
# 3. if match, open the two files, delete header data of temp_xlsx_date/*.xlsx
# 4. append date-data to temp_xlsx_main/*.xlsx
# 5. save the result to data\temp_xlsx\temp_xlsx_complete\*.xlsx

class _helper:

    @staticmethod
    def helper_get_ticker(ws: worksheet.Worksheet) -> str:
        return ws.title.split(" ")[-1]
    
    @staticmethod
    def get_temp_ws(date_path: Path) -> worksheet.Worksheet:
        temp_ws = pl.load_workbook(filename=date_path).active
        if not isinstance(temp_ws, worksheet.Worksheet):
            raise TypeError("Failed to load worksheet.")
        return temp_ws

def get_list_of_data() -> tuple[list[Path], list[Path]]:
    (list_main, list_date) = (
        sorted([f for f in Path(cp.TEMP_XLSX_MAIN_DIRPATH).glob("*.xlsx") if f.name[1:5] == cc.progress]),
        sorted([f for f in Path(cp.TEMP_XLSX_DATE_DIRPATH).glob("*.xlsx") if f.name[1:5] == cc.progress]),
    )
    return (list_main, list_date)

def compare_head_of_list(list_data_pair:tuple[Path, Path]) -> bool:
    (list_main, list_date) = (f.name for f in list_data_pair)
    if not list_main or not list_date:
        return False
    return list_main == list_date

def delete_header(temp_ws: worksheet.Worksheet) -> worksheet.Worksheet | None:
    if not isinstance(temp_ws, worksheet.Worksheet):
        return None
    for row in sorted(cc.REMOVE_ROW_FOR_XLSX2CSV, reverse=True):
        temp_ws.delete_rows(row)
    return temp_ws

def append_data(main_path: Path, date_ws: worksheet.Worksheet) -> worksheet.Worksheet | None:
    main_ws = pl.load_workbook(filename=main_path).active
    if not isinstance(main_ws, worksheet.Worksheet):
        raise ValueError("Failed to load main worksheet.")
    for row in date_ws.iter_rows(values_only=True):
        main_ws.append(row)
    return main_ws


def save_result(main_ws: worksheet.Worksheet) -> None:
    filename = _helper.helper_get_ticker(main_ws) + ".xlsx"
    save_path = Path(cp.TEMP_XLSX_COMPLETE_DIRPATH) / filename
    if isinstance(main_ws.parent, workbook.Workbook):
        main_ws.parent.save(save_path)
    else:
        raise ValueError("Worksheet does not have an associated workbook.")
    
def make_xlsx_shorter(xlsx_path: Path) -> None:
    wb = pl.load_workbook(filename=xlsx_path)
    ws = wb.active

    if not isinstance(ws, worksheet.Worksheet):
        raise ValueError("Failed to load worksheet.")

    inherited_ws_title = ws.title
    
    # get specific columns
    numpy_ws = np.array(list(ws.iter_rows(values_only=True)))
    serial_columns_temp = [list(col) for col in numpy_ws[:, :cc.REMOVE_COLUMN_FOR_XLSX2CSV[-1]].T]
    specific_columns = [list(col) for col in numpy_ws[:, cc.REMOVE_COLUMN_FOR_XLSX2CSV[-1]:].T]

    new_numpy_ws_temp = []

    serial_columns = serial_columns_temp[cc.SUBJECTNAME_COLUMN][1:cc.ACTUAL_RANGE_ROW["end"]]
    serial_columns.insert(1, "決算発表日")
    serial_columns[0] = "決算発表月"

    new_numpy_ws_temp.append(serial_columns) # append subject name column  # insert "決算発表月" column header
    # new_numpy_ws_temp.append(serial_columns[cc.SUBJECTCODE_COLUMN][cc.ACTUAL_RANGE_ROW["start"]:cc.ACTUAL_RANGE_ROW["end"]]) # append subject code column

    column_subjectname = serial_columns[cc.SUBJECTNAME_COLUMN]
    column_subjectcode = serial_columns[cc.SUBJECTCODE_COLUMN]

    # compare the column of 2 dates that are "決算短信" and "有価証券報告書"
    for column in specific_columns:
        if column[cc.KESSANTANSIN_DATE_ROW] is None and column[cc.YUUKAHOUKOKU_DATE_ROW] is None:
            raise ValueError(f"No date found for {column_subjectname[0]} ({column_subjectcode[0]}).")
        
        elif column[cc.KESSANTANSIN_DATE_ROW] is None:

            new_column = []
            # new_column.append(column[0]) # append header part
            new_column.append(column[1]) # append header part
            new_column.append(column[cc.YUUKAHOUKOKU_DATE_ROW]) # append date

            # and delete the "決算短信" date and related data
            for i in range(cc.YUUKAHOUKOKU_RANGE_ROW["start"], cc.YUUKAHOUKOKU_RANGE_ROW["end"]):
                new_column.append(column[i])
            
            # append them to the end of the column
            new_numpy_ws_temp.append(new_column)
        
        elif column[cc.YUUKAHOUKOKU_DATE_ROW] is None:
            new_column = []
            # new_column.append(column[0]) # append header part
            new_column.append(column[1]) # append header part
            new_column.append(column[cc.KESSANTANSIN_DATE_ROW]) # append date

            # and delete the "有価証券報告書" date and related data
            for i in range(cc.KESSANTANSIN_RANGE_ROW["start"], cc.KESSANTANSIN_RANGE_ROW["end"]):
                new_column.append(column[i])
            
            # append them to the end of the column
            new_numpy_ws_temp.append(new_column)

        # if same, prioritize the "有価証券報告書" date
        elif column[cc.KESSANTANSIN_DATE_ROW] == column[cc.YUUKAHOUKOKU_DATE_ROW]:
            new_column = []
            # new_column.append(column[0]) # append header part
            new_column.append(column[1]) # append header part
            new_column.append(column[cc.KESSANTANSIN_DATE_ROW]) # append date

            # and delete the "決算短信" date and related data
            for i in range(cc.YUUKAHOUKOKU_RANGE_ROW["start"], cc.YUUKAHOUKOKU_RANGE_ROW["end"]):
                new_column.append(column[i])
            
            # but if there is data in "決算短信" part, fill it in
            for i in range(cc.KESSANTANSIN_RANGE_ROW["start"], cc.KESSANTANSIN_RANGE_ROW["end"]):
                if new_column[i] == "-" and column[i] != "-":
                    new_column[i] = column[i]
            
            # append them to the end of the column
            new_numpy_ws_temp.append(new_column)

        # if different, first, extract the earlier date and related data
        elif column[cc.KESSANTANSIN_DATE_ROW] != column[cc.YUUKAHOUKOKU_DATE_ROW]:
            new_column_KESSANTANSIN = [] 
            new_column_YUUKAHOUKOKU = []
            # new_column_KESSANTANSIN.append(column[0]) # append header part
            new_column_KESSANTANSIN.append(column[1]) # append header part
            # new_column_YUUKAHOUKOKU.append(column[0]) # append header part
            new_column_YUUKAHOUKOKU.append(column[1]) # append header part
            new_column_KESSANTANSIN.append(column[cc.KESSANTANSIN_DATE_ROW]) # append date
            new_column_YUUKAHOUKOKU.append(column[cc.YUUKAHOUKOKU_DATE_ROW]) # append date

            # then, extract the later date and related data
            for i in range(cc.KESSANTANSIN_RANGE_ROW["start"], cc.KESSANTANSIN_RANGE_ROW["end"]):
                new_column_KESSANTANSIN.append(column[i])
            for i in range(cc.YUUKAHOUKOKU_RANGE_ROW["start"], cc.YUUKAHOUKOKU_RANGE_ROW["end"]):
                new_column_YUUKAHOUKOKU.append(column[i])
            
            # print(f"Different dates found for {column_subjectname[0]} ({column_subjectcode[0]}): {column[cc.KESSANTANSIN_DATE_ROW]} and {column[cc.YUUKAHOUKOKU_DATE_ROW]}")
            # append them to the end of the column by ascending order of date
            if new_column_KESSANTANSIN[1] == "-":
                new_numpy_ws_temp.append(new_column_YUUKAHOUKOKU)
            elif new_column_YUUKAHOUKOKU[1] == "-":
                new_numpy_ws_temp.append(new_column_KESSANTANSIN)
            elif dt.timestamp(dt.strptime(column[cc.KESSANTANSIN_DATE_ROW], "%Y/%m/%d")) < dt.timestamp(dt.strptime(column[cc.YUUKAHOUKOKU_DATE_ROW], "%Y/%m/%d")):
                new_numpy_ws_temp.append(new_column_KESSANTANSIN)
                new_numpy_ws_temp.append(new_column_YUUKAHOUKOKU)
            elif dt.timestamp(dt.strptime(column[cc.KESSANTANSIN_DATE_ROW], "%Y/%m/%d")) > dt.timestamp(dt.strptime(column[cc.YUUKAHOUKOKU_DATE_ROW], "%Y/%m/%d")):
                new_numpy_ws_temp.append(new_column_YUUKAHOUKOKU)
                new_numpy_ws_temp.append(new_column_KESSANTANSIN)
        
        else: # should not reach here
            raise ValueError("Unexpected condition encountered.")

    # save to excel file throgh pandas
    new_numpy_ws = np.column_stack(new_numpy_ws_temp)
    pd.DataFrame(new_numpy_ws).to_excel(xlsx_path, index=False, header=False, sheet_name=inherited_ws_title)


def main():
    list_main, list_date = get_list_of_data()

    for head_data_pair in zip(list_main, list_date):

        time.sleep(0.5)  # to avoid potential file access conflicts
        if not compare_head_of_list(head_data_pair):
            print("No matching files found.")
            return

        main_path, date_path = head_data_pair
        temp_ws = _helper.get_temp_ws(date_path)
        date_ws = delete_header(temp_ws)

        if date_ws is None:
            print("Failed to process date worksheet.")
            return
        
        main_ws = append_data(main_path, date_ws)

        if main_ws is None:
            print("Failed to process main worksheet.")
            return
        
        save_result(main_ws)

        make_xlsx_shorter(Path(cp.TEMP_XLSX_COMPLETE_DIRPATH) / (_helper.helper_get_ticker(main_ws) + ".xlsx"))

        print(f"Processed and saved: {main_ws.title}.xlsx")

if __name__ == "__main__":
    if debug := True:
        print(cp.TEMP_XLSX_DATE_DIRPATH)
        print(cp.TEMP_XLSX_MAIN_DIRPATH)
        print(cp.TEMP_XLSX_COMPLETE_DIRPATH)
    main()